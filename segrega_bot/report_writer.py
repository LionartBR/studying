# report_writer.py
from typing import List, Optional, Dict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os

def _autosize(ws):
    widths = {}
    for row in ws.iter_rows(values_only=True):
        for i, cell in enumerate(row, start=1):
            txt = "" if cell is None else str(cell)
            widths[i] = max(widths.get(i, 0), len(txt))
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, w + 2), 80)


def _append_manifest_sheet_from_rows(wb: Workbook, manifest_rows: Optional[List[Dict[str, str]]]):
    if not manifest_rows:
        return
    ws = wb.create_sheet("log")
    headers = ["source_path", "source_name", "collaborator", "created_path", "created_name", "status"]
    ws.append(headers)
    for r in manifest_rows:
        ws.append([
            r.get("source_path", ""),
            r.get("source_name", ""),
            r.get("collaborator", ""),
            r.get("created_path", ""),
            r.get("created_name", ""),
            r.get("status", ""),
        ])
    _autosize(ws)


def write_distribution_report(
    report_path: Optional[str],
    collaborators: List[str],
    rows: List[dict],
    not_found_collabs: List[str],
    files_no_match: List[str],
    manifest_rows: Optional[List[Dict[str, str]]] = None,  # <— agora recebe o manifest em memória
) -> Optional[str]:
    """
    rows: lista de dicts com:
      - collaborator: str
      - source_path: str
      - created_path: str  (pode ser "" quando houve match mas não criou destino)
    """
    if not report_path:
        return None

    os.makedirs(os.path.dirname(report_path) or ".", exist_ok=True)

    wb = Workbook()

    # Aba principal
    ws = wb.active
    ws.title = "Relatório de Distribuição"
    ws.append(["Colaborador", "Documento (origem)", "Arquivo criado", "Caminho do arquivo criado"])

    for r in rows:
        collab = r.get("collaborator", "")
        src = r.get("source_path", "")
        dst = r.get("created_path", "")

        src_name = os.path.basename(src) if src else ""
        created_name = os.path.basename(dst) if dst else "duplicata - ignorada"
        created_path_out = dst if dst else "-"

        ws.append([collab, src_name, created_name, created_path_out])

    for collab in not_found_collabs:
        ws.append([collab, "colaborador não localizado", "-", "-"])

    _autosize(ws)

    # Aba PDFs Sem Match
    ws2 = wb.create_sheet("PDFs Sem Match")
    ws2.append(["Nome do arquivo", "Local"])
    for p in sorted(files_no_match):
        ws2.append([os.path.basename(p), p])
    _autosize(ws2)

    # Aba manifest (em memória)
    _append_manifest_sheet_from_rows(wb, manifest_rows)

    wb.save(report_path)
    return report_path
